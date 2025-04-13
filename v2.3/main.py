import os
import sys
import json
import math
import pandas as pd
from fastapi import FastAPI, HTTPException, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from pydantic import BaseModel
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from io import BytesIO
from functools import lru_cache
from fastapi.staticfiles import StaticFiles

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:3000",
        "https://YOUR-NGROK-OR-RENDER-URL"
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["*"]
)

# --- Helper function to get paths relative to the script ---
def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        # In a standard run, __file__ is the script's path
        base_path = sys._MEIPASS
    except Exception:
        # Use the directory containing main.py as the base path
        base_path = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_path, relative_path)
# ---

def fix_duplicate_observaciones(df):
    cols = df.columns.tolist()
    if "OBSERVACIONES" in cols and "OBSERVACIONES\n" in cols:
        df = df.drop(columns=["OBSERVACIONES"])
    return df

def normalize_dataframe(df, required_cols):
    df = fix_duplicate_observaciones(df)
    for col in required_cols:
        if col not in df.columns:
            df[col] = ""
    df = df[required_cols]
    return df

# ==========================
# Define File Paths using resource_path
# ==========================
PROCEDURES_FILE = resource_path("maestro_procedimientos.xlsx")
DIAGNOSTICS_FILE = resource_path("maestro_diagnosticos.xlsx")
PROCESSED_MEDS_FILE = resource_path("medications_processed.csv")
DATA_FILE = resource_path("data.xlsx") # Also make data file relative

# ==========================
# Load Maestro Files
# ==========================
try:
    # Load procedures and diagnostics using defined paths
    if not os.path.exists(PROCEDURES_FILE):
        raise FileNotFoundError(f"Maestro file not found: {PROCEDURES_FILE}")
    proc_df = pd.read_excel(PROCEDURES_FILE)

    if not os.path.exists(DIAGNOSTICS_FILE):
        raise FileNotFoundError(f"Maestro file not found: {DIAGNOSTICS_FILE}")
    diag_df = pd.read_excel(DIAGNOSTICS_FILE)

    # Load the pre-processed medications CSV using defined path
    if not os.path.exists(PROCESSED_MEDS_FILE):
         raise FileNotFoundError(f"Processed medications file not found: {PROCESSED_MEDS_FILE}. Run preprocess_meds.py first.")
    med_df = pd.read_csv(PROCESSED_MEDS_FILE)

    # Ensure the required column 'concat_id' exists
    if 'concat_id' not in med_df.columns:
        raise HTTPException(status_code=500, detail=f"Column 'concat_id' not found in {PROCESSED_MEDS_FILE}.")
    # Ensure all values are strings and handle potential NaNs read from CSV
    med_df['concat_id'] = med_df['concat_id'].astype(str).fillna('')

except FileNotFoundError as e:
     # Raise HTTPException for file not found errors specifically
     raise HTTPException(status_code=500, detail=str(e))
except Exception as e:
    # General error loading files
    raise HTTPException(status_code=500, detail=f"Error loading maestro or processed files: {e}")


# DATA_FILE is now defined above using resource_path

REQUIRED_COLUMNS = [
    'CÓDIGO DEPENDENCIA\n(ESPECIALIDAD)\n',
    'PLANILLA',
    'FECHA ANTENCION',
    'TIPO DE BENEFICIARIO',
    'CEDULA',
    'NOMBRE DE BENEFICIARIO',
    'SEXO-GENERO',
    'FECHA DE NACIMIENTO BENEFICIERO',
    'EDAD BENEFICIERO',
    'TIPO DE SERVICIO/ATENCION',
    'CODIGO',
    'DESCRIPCIÓN',
    'DIAGNOSTICO PRINCIPAL CIE-10',
    'DIAGNSITICO SECUNDARIO 1',
    'DIAGNSITICO SECUNDARIO 2',
    'CANTIDAD',
    'VALOR UNITARIO',
    'DURACION CONSULTA',
    'PARENTESCO',
    'IDENTIFICACION AFILIADO',
    'NOMBRE AFIALIADO',
    'CODIGO DE DERIVACION',
    'NUMERO SECUNCIAL DERIVACION',
    'CONTINGENCIA CUBIERTA',
    'DIAGNOSTICO PRESUNTIVO O DIFINITIVO',
    'TIEMPO ANESTESIA',
    'DIAGNSITICO SECUNDARIO 3',
    'DIAGNSITICO SECUNDARIO 4',
    'DIAGNSITICO SECUNDARIO 5',
    'PORCENTAJE IVA',
    'VALOR IVA',
    'VALOR TOTAL',
    'GASTOS DE GESTIÓN (VALOR\nUNITARIO) / MODIFICADORES NO\nGEOGRÁFICOS (VALOR UNITARIO)',
    'FECHA DE INGRESO',
    'FECHA DE EGRESO',
    'MOTIVO DE EGRESO',
    'COBERTURA COMPARTIDA\n',
    'TIPO DE COBERTURA\n',
    'DISCAPACIDAD CERTIFICADA\n',
    'TIPO DE PRESTACIÓN\n',
    'TIPO DE MÉDICO',
    'FECHA AUTORIZADA PARA INICIO DE ATENCIÓN \n',
    'OBSERVACIONES\n'
]

if os.path.exists(DATA_FILE):
    df = pd.read_excel(DATA_FILE)
    df.columns = df.columns.str.strip()
    df = normalize_dataframe(df, REQUIRED_COLUMNS)
else:
    df = pd.DataFrame(columns=REQUIRED_COLUMNS)

grid_columns = [
    'FECHA DE INGRESO',
    'FECHA DE EGRESO',
    'CÓDIGO DEPENDENCIA (ESPECIALIDAD)',
    'FECHA ANTENCION',
    'CEDULA',
    'NOMBRE DE BENEFICIARIO',
    'CODIGO',
    'DESCRIPCIÓN',
    'DIAGNOSTICO PRINCIPAL CIE-10',
    'DIAGNSITICO SECUNDARIO 1',
    'CANTIDAD',
    'DIAGNOSTICO PRESUNTIVO O DIFINITIVO',
    'OBSERVACIONES'
]

color_fills = [
    PatternFill(start_color="FF92D050", end_color="FF92D050", fill_type="darkGrid"),
    PatternFill(start_color="FF00B0F0", end_color="FF00B0F0", fill_type="darkTrellis"),
    PatternFill(start_color="FFFFC000", end_color="FFFFC000", fill_type="lightGrid"),
    PatternFill(start_color="FF7030A0", end_color="FF7030A0", fill_type="lightTrellis"),
    PatternFill(start_color="FF00B050", end_color="FF00B050", fill_type="darkHorizontal"),
    PatternFill(start_color="FFED7D31", end_color="FFED7D31", fill_type="darkVertical"),
    PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="lightHorizontal"),
    PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="lightVertical"),
    PatternFill(start_color="FFBFBFBF", end_color="FFBFBFBF", fill_type="darkDown"),
    PatternFill(start_color="FFFF00FF", end_color="FFFF00FF", fill_type="darkUp")
]

# -----------------------------
# Pydantic Models
# -----------------------------
class EntryItem(BaseModel):
    name: str = ""
    code: str = ""
    quantity: int = 0

class NewEntry(BaseModel):
    paciente: str = ""
    diagnostico_name: str = ""
    diagnostico_code: str = ""
    diagnostico_secundario_name: str = ""
    diagnostico_secundario_code: str = ""
    fecha_ingreso: str = ""
    fecha_egreso: str = ""
    procedimientos: list[EntryItem] = []
    medicamentos: list[EntryItem] = []
    insumos: list[EntryItem] = []

class DeleteRows(BaseModel):
    ids: list[int]

# -----------------------------
# Endpoints
# -----------------------------
@app.post("/upload/")
async def upload_file(file: UploadFile = File(...)):
    global df
    try:
        contents = await file.read()
        file_location = f"./{file.filename}"
        with open(file_location, "wb") as f:
            f.write(contents)
        if file.filename.lower().endswith('.csv'):
            temp_df = pd.read_csv(BytesIO(contents))
        else:
            temp_df = pd.read_excel(BytesIO(contents))
        temp_df.columns = temp_df.columns.str.strip()
        temp_df = normalize_dataframe(temp_df, REQUIRED_COLUMNS)
        df = temp_df
        return {"message": "File uploaded and loaded successfully."}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/data/")
def get_data():
    records = df.to_dict(orient="records")
    for idx, record in enumerate(records):
        record["id"] = idx
        for key, value in record.items():
            if isinstance(value, float) and math.isnan(value):
                record[key] = None
    return records

@app.get("/sync/diagnostic/")
def sync_diagnostic(name: str = None, code: str = None):
    if name:
        row = diag_df[diag_df["NOMBRE"].str.lower() == name.lower()]
    elif code:
        row = diag_df[diag_df["CÓDIGO"].astype(str) == str(code)]
    else:
        raise HTTPException(status_code=400, detail="Provide either name or code")
    if row.empty:
        raise HTTPException(status_code=404, detail="Diagnostic not found")
    return {"name": row.iloc[0]["NOMBRE"], "code": row.iloc[0]["CÓDIGO"]}

@app.get("/search/diagnostics/")
def search_diagnostics(query: str):
    results = diag_df[diag_df["NOMBRE"].str.contains(query, case=False, na=False)].head(50)
    return results[["NOMBRE", "CÓDIGO"]].to_dict(orient="records")

@app.get("/search/diagnostics/code/")
def search_diagnostics_code(query: str):
    mask = diag_df["CÓDIGO"].astype(str).str.contains(query, case=False, na=False)
    results = diag_df[mask].head(50)
    return results[["NOMBRE", "CÓDIGO"]].to_dict(orient="records")

@app.get("/search/procedures/")
def search_procedures(query: str):
    results = proc_df[proc_df["DESCRIPCIÓN"].str.contains(query, case=False, na=False)].head(50)
    return results[["DESCRIPCIÓN", "CÓDIGO"]].to_dict(orient="records")

@app.get("/search/medications/")
def search_medications(query: str):
    # Search in the 'concat_id' column from the processed CSV
    results = med_df[med_df["concat_id"].str.contains(query, case=False, na=False)].head(50)
    # Return 'concat_id' for both display (label) and identifier ('code')
    out = []
    for _, row in results.iterrows():
        concat_val = str(row["concat_id"])
        out.append({"label": concat_val, "value": concat_val, "code": concat_val})
    return out

@lru_cache(maxsize=1)
@app.get("/patients/full/")
def get_patients_full():
    return sorted(df["NOMBRE DE BENEFICIARIO"].dropna().unique().tolist())

@lru_cache(maxsize=1)
@app.get("/medications/full/")
def get_medications_full():
    # Return 'concat_id' for both display (label) and identifier ('code')
    out = []
    for _, row in med_df.iterrows():
         concat_val = str(row["concat_id"])
         out.append({"label": concat_val, "value": concat_val, "code": concat_val})
    return out

@lru_cache(maxsize=1)
@app.get("/procedures/full/")
def get_procedures_full():
    out = []
    for _, row in proc_df.iterrows():
        out.append({
            "DESCRIPCIÓN": row["DESCRIPCIÓN"],
            "CÓDIGO": str(row["CÓDIGO"])
        })
    return out

@lru_cache(maxsize=1)
@app.get("/diagnostics/full/")
def get_diagnostics_full():
    out = []
    for _, row in diag_df.iterrows():
        out.append({
            "NOMBRE": row["NOMBRE"],
            "CÓDIGO": str(row["CÓDIGO"])
        })
    return out

@app.get("/download/")
def download_file():
    if not os.path.exists(DATA_FILE):
        raise HTTPException(status_code=404, detail="Data file not found")
    return FileResponse(
        path=DATA_FILE,
        filename="data.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.post("/add/")
def add_entry(entry: NewEntry):
    global df
    base_row = {
        "NOMBRE DE BENEFICIARIO": entry.paciente,
        "DIAGNOSTICO PRINCIPAL CIE-10": entry.diagnostico_code,
        "DIAGNOSTICO PRESUNTIVO O DIFINITIVO": entry.diagnostico_name,
        "DIAGNSITICO SECUNDARIO 1": entry.diagnostico_secundario_code,
        "FECHA DE INGRESO": entry.fecha_ingreso,
        "FECHA DE EGRESO": entry.fecha_egreso,
        "OBSERVACIONES": ""
    }
    new_entries = []
    for item in entry.procedimientos:
        if item.name:
            row = base_row.copy()
            row["DESCRIPCIÓN"] = item.name
            row["CODIGO"] = item.code # Procedures still use CODIGO
            row["CANTIDAD"] = item.quantity
            new_entries.append(row)
    for item in entry.medicamentos:
        if item.name: # item.name is the concat_id string
            row = base_row.copy()
            # Store the concat_id string in DESCRIPCIÓN
            # Leave CODIGO blank for medications as requested
            row["DESCRIPCIÓN"] = item.name
            row["CODIGO"] = "" # Leave code blank for medications
            row["CANTIDAD"] = item.quantity
            new_entries.append(row)
    for item in entry.insumos: # Assuming insumos might be similar or needs separate handling
        if item.name:
            row = base_row.copy()
            row["DESCRIPCIÓN"] = item.name
            row["CODIGO"] = item.code
            row["CANTIDAD"] = item.quantity
            new_entries.append(row)
    current_rows = df.to_dict(orient="records")
    for new_row in new_entries:
        patient = new_row["NOMBRE DE BENEFICIARIO"]
        insertion_index = None
        for i in range(len(current_rows) - 1, -1, -1):
            if current_rows[i]["NOMBRE DE BENEFICIARIO"] == patient:
                # Copy relevant fields from the found existing row
                new_row["CÓDIGO DEPENDENCIA\n(ESPECIALIDAD)\n"] = current_rows[i].get("CÓDIGO DEPENDENCIA\n(ESPECIALIDAD)\n", "")
                new_row["FECHA ANTENCION"] = current_rows[i].get("FECHA ANTENCION", "") # Keep original date? Or use new entry date? Currently uses existing.
                new_row["CEDULA"] = current_rows[i].get("CEDULA", "")
                # --- Add copying for other required fields ---
                new_row["SEXO-GENERO"] = current_rows[i].get("SEXO-GENERO", "")
                new_row["PLANILLA"] = current_rows[i].get("PLANILLA", "")
                new_row["TIPO DE BENEFICIARIO"] = current_rows[i].get("TIPO DE BENEFICIARIO", "")
                new_row["PARENTESCO"] = current_rows[i].get("PARENTESCO", "")
                new_row["IDENTIFICACION AFILIADO"] = current_rows[i].get("IDENTIFICACION AFILIADO", "")
                new_row["NOMBRE AFIALIADO"] = current_rows[i].get("NOMBRE AFIALIADO", "")
                # Copy other potentially relevant fields if needed, e.g., FECHA DE NACIMIENTO, EDAD?
                new_row["FECHA DE NACIMIENTO BENEFICIERO"] = current_rows[i].get("FECHA DE NACIMIENTO BENEFICIERO", "")
                new_row["EDAD BENEFICIERO"] = current_rows[i].get("EDAD BENEFICIERO", "")
                # --- Add copying for even more potentially relevant static fields ---
                new_row["TIPO DE SERVICIO/ATENCION"] = current_rows[i].get("TIPO DE SERVICIO/ATENCION", "")
                new_row["DURACION CONSULTA"] = current_rows[i].get("DURACION CONSULTA", "")
                new_row["CODIGO DE DERIVACION"] = current_rows[i].get("CODIGO DE DERIVACION", "")
                new_row["NUMERO SECUNCIAL DERIVACION"] = current_rows[i].get("NUMERO SECUNCIAL DERIVACION", "")
                new_row["CONTINGENCIA CUBIERTA"] = current_rows[i].get("CONTINGENCIA CUBIERTA", "")
                new_row["TIEMPO ANESTESIA"] = current_rows[i].get("TIEMPO ANESTESIA", "") # If applicable
                new_row["MOTIVO DE EGRESO"] = current_rows[i].get("MOTIVO DE EGRESO", "") # If applicable
                new_row["COBERTURA COMPARTIDA\n"] = current_rows[i].get("COBERTURA COMPARTIDA\n", "")
                new_row["TIPO DE COBERTURA\n"] = current_rows[i].get("TIPO DE COBERTURA\n", "")
                new_row["DISCAPACIDAD CERTIFICADA\n"] = current_rows[i].get("DISCAPACIDAD CERTIFICADA\n", "")
                new_row["TIPO DE PRESTACIÓN\n"] = current_rows[i].get("TIPO DE PRESTACIÓN\n", "")
                new_row["TIPO DE MÉDICO"] = current_rows[i].get("TIPO DE MÉDICO", "")
                new_row["FECHA AUTORIZADA PARA INICIO DE ATENCIÓN \n"] = current_rows[i].get("FECHA AUTORIZADA PARA INICIO DE ATENCIÓN \n", "")
                # --- End added fields ---
                insertion_index = i + 1
                break
        if insertion_index is None:
            # Handle case where patient is entirely new?
            # Currently appends to end without copying any details.
            # Might need default values or error handling if patient must exist.
            current_rows.append(new_row)
        else:
            current_rows.insert(insertion_index, new_row)
    df = pd.DataFrame(current_rows)
    # Save and load workbook using the defined DATA_FILE path
    df.to_excel(DATA_FILE, index=False, columns=df.columns.tolist())
    try:
        wb = load_workbook(DATA_FILE)
        ws = wb.active
        # Ensure header reading handles potential None values gracefully
        header = [str(cell.value) if cell.value is not None else '' for cell in ws[1]]
        patient_idx = header.index("NOMBRE DE BENEFICIARIO") + 1
        date_idx = header.index("FECHA ANTENCION") + 1
        current_color_index = -1
        prev_key = None
        for r in range(2, ws.max_row + 1):
            patient_val = ws.cell(row=r, column=patient_idx).value
            date_val = ws.cell(row=r, column=date_idx).value
            key = (date_val, patient_val)
            if key != prev_key:
                current_color_index = (current_color_index + 1) % len(color_fills)
                prev_key = key
            fill_to_use = color_fills[current_color_index]
            for c in range(1, ws.max_column + 1):
                ws.cell(row=r, column=c).fill = fill_to_use
        wb.save(DATA_FILE)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error saving colored file: {e}")
    return {"message": "Entry added successfully!"}

@app.post("/delete/")
def delete_rows(delete_request: DeleteRows):
    global df
    records = df.to_dict(orient="records")
    new_records = [record for idx, record in enumerate(records) if idx not in delete_request.ids]
    df = pd.DataFrame(new_records)
    # Save using the defined DATA_FILE path
    df.to_excel(DATA_FILE, index=False, columns=df.columns.tolist())
    return {"message": "Filas eliminadas exitosamente."}

# @app.post("/save/") # This endpoint seems redundant if /add/ already saves
# def save_file():
#     df.to_excel(DATA_FILE, index=False, columns=df.columns.tolist())
#     return {"message": "File saved successfully."}

@app.on_event("shutdown")
def save_state():
    # Consider saving the current state of df if needed on shutdown
    # df.to_excel(DATA_FILE, index=False, columns=df.columns.tolist())
    pass

# Serve static files from frontend build directory, relative to this script
build_path = resource_path(os.path.join("frontend", "build"))
if os.path.exists(build_path):
    app.mount("/", StaticFiles(directory=build_path, html=True), name="static")
else:
    print(f"Warning: React build directory not found at {build_path}. Frontend will not be served.")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
