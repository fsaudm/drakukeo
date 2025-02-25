import sys
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import pandas as pd
import os
import json

STATE_FILENAME = "state.json"

# --------------------------
# Load maestro files
# --------------------------
try:
    proc_df = pd.read_excel("maestro_procedimientos.xlsx")
    med_df = pd.read_excel("maestro_medicamentos.xlsx")
    diag_df = pd.read_excel("maestro_diagnosticos.xlsx")
except Exception as e:
    messagebox.showerror("Error", f"Error loading maestro files: {e}")
    exit()

# Prepare concatenated field for medications
med_df["concat"] = med_df["DESCRIPCIÓN"].astype(str) + " " + med_df["PRESENTACION"].astype(str)

# --------------------------
# Global variables
# --------------------------
main_file_path = None
required_columns = [
    'CÓDIGO DEPENDENCIA\n(ESPECIALIDAD)\n',
    'PLANILLA',
    'FECHA ANTENCION',
    'TIPO DE BENEFICIARIO',
    'CEDULA',
    'NOMBRE DE BENEFICIARIO',
    'SEXO-GENERO',
    'FECHA DE NACIMIENTO BENEFICIERO',
    'EDAD BENEFICIERO',
    'TIPO DE SERVICIO/ATENCION',
    'CODIGO',
    'DESCRIPCIÓN',
    'OBSERVACIONES',
    'DIAGNOSTICO PRINCIPAL CIE-10',
    'DIAGNSITICO SECUNDARIO 1',
    'DIAGNSITICO SECUNDARIO 2',
    'CANTIDAD',
    'VALOR UNITARIO',
    'DURACION CONSULTA',
    'PARENTESCO',
    'IDENTIFICACION AFILIADO',
    'NOMBRE AFIALIADO',
    'CODIGO DE DERIVACION',
    'NUMERO SECUNCIAL DERIVACION',
    'CONTINGENCIA CUBIERTA',
    'DIAGNOSTICO PRESUNTIVO O DIFINITIVO',
    'TIEMPO ANESTESIA',
    'DIAGNSITICO SECUNDARIO 3',
    'DIAGNSITICO SECUNDARIO 4',
    'DIAGNSITICO SECUNDARIO 5',
    'PORCENTAJE IVA',
    'VALOR IVA',
    'VALOR TOTAL',
    'GASTOS DE GESTIÓN (VALOR\nUNITARIO) / MODIFICADORES NO\nGEOGRÁFICOS (VALOR UNITARIO)',
    'FECHA DE INGRESO',
    'FECHA DE EGRESO',
    'MOTIVO DE EGRESO',
    'COBERTURA COMPARTIDA\n',
    'TIPO DE COBERTURA\n',
    'DISCAPACIDAD CERTIFICADA\n',
    'TIPO DE PRESTACIÓN\n',
    'TIPO DE MÉDICO',
    'FECHA AUTORIZADA PARA INICIO DE ATENCIÓN \n',
    'OBSERVACIONES\n',
    'MARCA FINAL (SIEMPRE F)'
]
main_df = pd.DataFrame(columns=required_columns)

# The grid view will only show these columns:
grid_columns = [
    'CÓDIGO DEPENDENCIA\n(ESPECIALIDAD)\n',
    'FECHA ANTENCION',
    'CEDULA',
    'NOMBRE DE BENEFICIARIO',
    'CODIGO',
    'DESCRIPCIÓN',
    'OBSERVACIONES',
    'DIAGNOSTICO PRINCIPAL CIE-10',
    'CANTIDAD',
    'DIAGNOSTICO PRESUNTIVO O DIFINITIVO',
    'OBSERVACIONES\n'
]

# --------------------------
# Utility: Clear field on Escape
# --------------------------
def clear_on_escape(event):
    try:
        event.widget.delete(0, tk.END)
    except Exception:
        pass

# --------------------------
# Function to load main file (with column check)
# --------------------------
def load_main_file():
    global main_df, main_file_path, patients
    file_path = filedialog.askopenfilename(
        title="Seleccionar archivo main",
        filetypes=[("Excel Files", "*.xlsx"), ("CSV Files", "*.csv")]
    )
    if file_path:
        load_main_file_state(file_path)

def load_main_file_state(file_path):
    global main_df, main_file_path, patients
    try:
        ext = os.path.splitext(file_path)[1].lower()
        if ext == '.csv':
            df = pd.read_csv(file_path)
        else:
            df = pd.read_excel(file_path)
        missing = [col for col in required_columns if col not in df.columns]
        if missing:
            messagebox.showerror("Error",
                f"El archivo seleccionado no tiene las columnas requeridas.\nSe esperaban:\n{', '.join(required_columns)}\nFaltan:\n{', '.join(missing)}")
            return
        main_df = df
        main_file_path = file_path
        patients = main_df["NOMBRE DE BENEFICIARIO"].dropna().unique().tolist()
        paciente_combo.set_completion_list(patients)
        refresh_treeview()
    except Exception as e:
        messagebox.showerror("Error", f"Error al cargar el archivo: {e}")

# --------------------------
# Autocomplete Combobox class
# --------------------------
class AutocompleteCombobox(ttk.Combobox):
    def __init__(self, master=None, substring_search=False, **kw):
        super().__init__(master, **kw)
        self.substring_search = substring_search
        self.bind('<KeyRelease>', self._handle_keyrelease)
        self.bind("<Escape>", clear_on_escape)
    def set_completion_list(self, completion_list):
        self._completion_list = sorted(completion_list, key=str.lower)
        self['values'] = self._completion_list
    def _handle_keyrelease(self, event):
        if event.keysym in ("Return", "Tab", "BackSpace", "Left", "Right", "Up", "Down", "Escape"):
            return
        value = self.get().lower()
        if value == "":
            filtered = self._completion_list
        else:
            if self.substring_search:
                filtered = [item for item in self._completion_list if value in item.lower()]
            else:
                filtered = [item for item in self._completion_list if item.lower().startswith(value)]
        self['values'] = filtered if filtered else self._completion_list
        self.after(2000, lambda: self.event_generate('<Down>'))

# --------------------------
# Prepare dropdown lists
# --------------------------
patients = []  # updated on file load
diag_values = diag_df["NOMBRE"].dropna().unique().tolist()
proc_list = proc_df["DESCRIPCIÓN"].dropna().unique().tolist()
med_list = med_df["concat"].dropna().unique().tolist()

# --------------------------
# Create main window and layout
# --------------------------
root = tk.Tk()
root.title("Registro de Servicios de Pacientes")
root.rowconfigure(0, weight=1)
root.columnconfigure(0, weight=1)

# Force light theme on macOS
if sys.platform == "darwin":
    try:
        root.tk.call("tk::mac::setTheme", "Aqua")
    except Exception as e:
        print("Could not force light theme:", e)

paned = ttk.Panedwindow(root, orient=tk.HORIZONTAL)
paned.grid(row=0, column=0, sticky="nsew")
left_frame = ttk.Frame(paned, width=300)
right_frame = ttk.Frame(paned)
paned.add(left_frame, weight=0)
paned.add(right_frame, weight=1)

# --------------------------
# Left Frame Widgets
# --------------------------
ttk.Button(left_frame, text="Seleccionar Archivo Main", command=load_main_file)\
    .grid(row=0, column=0, columnspan=4, pady=5)

ttk.Label(left_frame, text="Seleccione Paciente:").grid(row=1, column=0, padx=5, pady=5, sticky="w")
paciente_combo = AutocompleteCombobox(left_frame)
paciente_combo.set_completion_list(patients)
paciente_combo.grid(row=1, column=1, padx=5, pady=5)

ttk.Label(left_frame, text="Observaciones:").grid(row=2, column=0, padx=5, pady=5, sticky="w")
obs_entry = ttk.Entry(left_frame, width=40)
obs_entry.grid(row=2, column=1, padx=5, pady=5)

# Diagnóstico: name and code fields
ttk.Label(left_frame, text="Seleccione Diagnóstico:").grid(row=3, column=0, padx=5, pady=5, sticky="w")
diag_combo = AutocompleteCombobox(left_frame)
diag_combo.set_completion_list(diag_values)
diag_combo.grid(row=3, column=1, padx=5, pady=5)
ttk.Label(left_frame, text="Código Diagnóstico:").grid(row=3, column=2, padx=5, pady=5, sticky="w")
diag_code_entry = ttk.Entry(left_frame, width=15)
diag_code_entry.grid(row=3, column=3, padx=5, pady=5)
diag_code_entry.bind("<Escape>", clear_on_escape)
def update_diag_from_code(event=None):
    code = diag_code_entry.get().strip()
    if code:
        row = diag_df[diag_df["CÓDIGO"] == code]
        if not row.empty:
            name = row.iloc[0]["NOMBRE"]
            diag_combo.set(name)
diag_code_entry.bind("<KeyRelease>", update_diag_from_code)
def update_diag_code(event):
    selected_diag = diag_combo.get().strip()
    if selected_diag:
        row_diag = diag_df[diag_df["NOMBRE"] == selected_diag]
        if not row_diag.empty:
            code = row_diag.iloc[0]["CÓDIGO"]
            diag_code_entry.delete(0, tk.END)
            diag_code_entry.insert(0, code)
diag_combo.bind("<<ComboboxSelected>>", update_diag_code)

# --------------------------
# Service Frame: Procedimientos and Medicamentos
# --------------------------
service_frame = ttk.Frame(left_frame)
service_frame.grid(row=4, column=0, columnspan=4, padx=5, pady=5)

# Procedimientos: combobox + "Cantidad"
proc_frame = ttk.LabelFrame(service_frame, text="Procedimientos")
proc_frame.grid(row=0, column=0, padx=5, pady=5)
proc_entries = []
proc_qty_entries = []
for i in range(5):
    ttk.Label(proc_frame, text=f"Procedimiento {i+1}:").grid(row=i, column=0, padx=3, pady=2, sticky="w")
    ac = AutocompleteCombobox(proc_frame)
    ac.set_completion_list(proc_list)
    ac.grid(row=i, column=1, padx=3, pady=2)
    proc_entries.append(ac)
    ttk.Label(proc_frame, text="Cantidad:").grid(row=i, column=2, padx=3, pady=2, sticky="w")
    qty = ttk.Entry(proc_frame, width=8)
    qty.grid(row=i, column=3, padx=3, pady=2)
    proc_qty_entries.append(qty)

# Medicamentos: combobox with substring search + "Cantidad"
med_frame = ttk.LabelFrame(service_frame, text="Medicamentos")
med_frame.grid(row=0, column=1, padx=5, pady=5)
med_entries = []
med_qty_entries = []
for i in range(5):
    ttk.Label(med_frame, text=f"Medicamento {i+1}:").grid(row=i, column=0, padx=3, pady=2, sticky="w")
    ac = AutocompleteCombobox(med_frame, substring_search=True)
    ac.set_completion_list(med_list)
    ac.grid(row=i, column=1, padx=3, pady=2)
    med_entries.append(ac)
    ttk.Label(med_frame, text="Cantidad:").grid(row=i, column=2, padx=3, pady=2, sticky="w")
    qty = ttk.Entry(med_frame, width=8)
    qty.grid(row=i, column=3, padx=3, pady=2)
    med_qty_entries.append(qty)

# Insumos: free text + "Cantidad" (no code lookup)
insumos_frame = ttk.LabelFrame(left_frame, text="Insumos")
insumos_frame.grid(row=5, column=0, columnspan=4, padx=5, pady=5)
insumo_desc_entries = []
insumo_qty_entries = []
for i in range(5):
    ttk.Label(insumos_frame, text=f"Insumo {i+1}:").grid(row=i, column=0, padx=3, pady=2, sticky="w")
    desc = ttk.Entry(insumos_frame, width=30)
    desc.grid(row=i, column=1, padx=3, pady=2)
    insumo_desc_entries.append(desc)
    ttk.Label(insumos_frame, text="Cantidad:").grid(row=i, column=2, padx=3, pady=2, sticky="w")
    qty = ttk.Entry(insumos_frame, width=8)
    qty.grid(row=i, column=3, padx=3, pady=2)
    insumo_qty_entries.append(qty)

ttk.Button(left_frame, text="Agregar Entrada", command=lambda: add_entry())\
    .grid(row=6, column=1, pady=10)

# --------------------------
# Right Frame: Treeview (Grid Visualizer)
# --------------------------
tree = ttk.Treeview(right_frame, columns=grid_columns, show="headings")
for col in grid_columns:
    tree.heading(col, text=col)
    tree.column(col, width=120, minwidth=80, stretch=True)
tree.grid(row=0, column=0, sticky="nsew")
vsb = ttk.Scrollbar(right_frame, orient="vertical", command=tree.yview)
hsb = ttk.Scrollbar(right_frame, orient="horizontal", command=tree.xview)
tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
vsb.grid(row=0, column=1, sticky="ns")
hsb.grid(row=1, column=0, sticky="ew")
right_frame.rowconfigure(0, weight=1)
right_frame.columnconfigure(0, weight=1)

def refresh_treeview():
    tree.delete(*tree.get_children())
    for idx, (_, row) in enumerate(main_df.iterrows()):
        tag = 'even' if idx % 2 == 0 else 'odd'
        values = [row.get(col, "") for col in grid_columns]
        tree.insert("", "end", values=values, tags=(tag,))
    tree.tag_configure('even', background='white')
    tree.tag_configure('odd', background='#f0f0ff')

def add_entry():
    global main_df
    selected_patient = paciente_combo.get().strip()
    if not selected_patient:
        messagebox.showwarning("Warning", "Seleccione un paciente.")
        return
    mask = main_df["NOMBRE DE BENEFICIARIO"].str.strip().str.lower() == selected_patient.lower()
    if not mask.any():
        messagebox.showerror("Error", "Paciente no encontrado en el archivo main.")
        return
    base_row = main_df[mask].iloc[0].to_dict()
    base_row["TIPO DE SERVICIO/ATENCION"] = "EMERGENCIA"
    base_row["OBSERVACIONES"] = obs_entry.get()
    selected_diag = diag_combo.get().strip()
    diag_code_val = diag_code_entry.get().strip()
    if selected_diag:
        row_diag = diag_df[diag_df["NOMBRE"] == selected_diag]
        if row_diag.empty and diag_code_val:
            row_diag = diag_df[diag_df["CÓDIGO"] == diag_code_val]
        if row_diag.empty:
            messagebox.showerror("Error", "No se encontró un diagnóstico válido.")
            return
        principal_code = row_diag.iloc[0]["CÓDIGO"]
        base_row["DIAGNOSTICO PRINCIPAL CIE-10"] = principal_code
    elif diag_code_val:
        row_diag = diag_df[diag_df["CÓDIGO"] == diag_code_val]
        if row_diag.empty:
            messagebox.showerror("Error", "Código de diagnóstico no encontrado.")
            return
        principal_code = row_diag.iloc[0]["CÓDIGO"]
        base_row["DIAGNOSTICO PRINCIPAL CIE-10"] = principal_code
        diag_combo.set(row_diag.iloc[0]["NOMBRE"])
    else:
        messagebox.showwarning("Warning", "Seleccione o ingrese un diagnóstico.")
        return
    base_row["DIAGNOSTICO PRESUNTIVO O DIFINITIVO"] = diag_code_val

    new_rows = []
    for i, entry in enumerate(proc_entries):
        text = entry.get().strip()
        qty = proc_qty_entries[i].get().strip()
        if text:
            code_row = proc_df[proc_df["DESCRIPCIÓN"].str.lower() == text.lower()]
            if code_row.empty:
                messagebox.showerror("Error", f"No se encontró el código para el procedimiento: {text}")
                continue
            selected_code = code_row.iloc[0]["CÓDIGO"]
            new_row = base_row.copy()
            new_row["CODIGO"] = selected_code
            new_row["DESCRIPCIÓN"] = text
            new_row["CANTIDAD"] = qty
            new_rows.append(new_row)
    for i, entry in enumerate(med_entries):
        text = entry.get().strip()
        qty = med_qty_entries[i].get().strip()
        if text:
            code_row = med_df[med_df["concat"].str.lower() == text.lower()]
            if code_row.empty:
                messagebox.showerror("Error", f"No se encontró el código para el medicamento: {text}")
                continue
            selected_code = code_row.iloc[0]["CÓDIGO"]
            new_row = base_row.copy()
            new_row["CODIGO"] = selected_code
            new_row["DESCRIPCIÓN"] = text
            new_row["CANTIDAD"] = qty
            new_rows.append(new_row)
    for i, desc_entry in enumerate(insumo_desc_entries):
        desc = desc_entry.get().strip()
        qty = insumo_qty_entries[i].get().strip()
        if desc:
            new_row = base_row.copy()
            new_row["CODIGO"] = ""
            new_row["DESCRIPCIÓN"] = desc
            new_row["CANTIDAD"] = qty
            new_rows.append(new_row)
    if not new_rows:
        messagebox.showwarning("Warning", "No se ingresó ningún procedimiento, medicamento o insumo.")
        return
    new_df = pd.DataFrame(new_rows)
    matching_indices = main_df.index[
        main_df["NOMBRE DE BENEFICIARIO"].str.strip().str.lower() == selected_patient.lower()
    ].tolist()
    if matching_indices:
        insert_index = max(matching_indices) + 1
    else:
        insert_index = len(main_df)
    part1 = main_df.iloc[:insert_index]
    part2 = main_df.iloc[insert_index:]
    main_df = pd.concat([part1, new_df, part2], ignore_index=True)
    
    try:
        if main_file_path and main_file_path.endswith('.csv'):
            main_df.to_csv(main_file_path, index=False)
        else:
            filename = main_file_path or "main.xlsx"
            # Save using the same columns as read (unchanged)
            main_df.to_excel(filename, index=False, columns=main_df.columns.tolist())
            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill
            wb = load_workbook(filename)
            ws = wb.active
            header = [cell.value for cell in ws[1]]
            try:
                patient_idx = header.index("NOMBRE DE BENEFICIARIO") + 1
                date_idx = header.index("FECHA ANTENCION") + 1
            except ValueError:
                wb.save(filename)
            fill1 = PatternFill(start_color="FF92D050", end_color="FF92D050", fill_type="solid")
            fill2 = PatternFill(start_color="FF00B0F0", end_color="FF00B0F0", fill_type="solid")
            current_fill = fill1
            prev_key = None
            for r in range(2, ws.max_row + 1):
                patient_val = ws.cell(row=r, column=patient_idx).value
                date_val = ws.cell(row=r, column=date_idx).value
                key = (patient_val, date_val)
                if key != prev_key:
                    if prev_key is not None:
                        current_fill = fill2 if current_fill == fill1 else fill1
                    prev_key = key
                for c in range(1, ws.max_column + 1):
                    ws.cell(row=r, column=c).fill = current_fill
            wb.save(filename)
    except Exception as e:
        messagebox.showerror("Error", f"Error al guardar el archivo: {e}")
        return
    messagebox.showinfo("Success", "Entrada agregada correctamente.")
    obs_entry.delete(0, tk.END)
    diag_combo.set("")
    diag_code_entry.delete(0, tk.END)
    for entry in proc_entries:
        entry.delete(0, tk.END)
    for qty in proc_qty_entries:
        qty.delete(0, tk.END)
    for entry in med_entries:
        entry.delete(0, tk.END)
    for qty in med_qty_entries:
        qty.delete(0, tk.END)
    for entry in insumo_desc_entries:
        entry.delete(0, tk.END)
    for qty in insumo_qty_entries:
        qty.delete(0, tk.END)
    refresh_treeview()

def on_tree_select(event):
    selected = tree.selection()
    if selected:
        item = tree.item(selected[0])
        values = item.get("values", [])
        if len(values) >= 4:
            paciente_combo.set(values[3])
        if len(values) >= 8 and values[7]:
            diag_code_entry.delete(0, tk.END)
            diag_code_entry.insert(0, values[7])
            update_diag_from_code()

tree.bind("<<TreeviewSelect>>", on_tree_select)

# --------------------------
# State saving/loading functions
# --------------------------
def save_state():
    state = {
        "main_file_path": main_file_path,
        "vertical_scroll": vsb.get(),
        "tree_selection": tree.selection()
    }
    with open(STATE_FILENAME, "w") as f:
        json.dump(state, f)

def load_state():
    if os.path.exists(STATE_FILENAME):
        with open(STATE_FILENAME, "r") as f:
            state = json.load(f)
        file_path = state.get("main_file_path", "")
        vertical_scroll = state.get("vertical_scroll", None)
        selection = state.get("tree_selection", [])
        if file_path and os.path.exists(file_path):
            load_main_file_state(file_path)
            tree_items = tree.get_children()
            valid_selection = [item for item in selection if item in tree_items]
            if valid_selection:
                tree.selection_set(valid_selection)
                tree.see(valid_selection[0])
            if vertical_scroll is not None:
                tree.yview_moveto(vertical_scroll[0])

load_state()

def on_closing():
    save_state()
    root.destroy()

root.protocol("WM_DELETE_WINDOW", on_closing)

# --------------------------
# Run application
# --------------------------
if __name__ == '__main__':
    root.mainloop()
