import os
import sys
import json
import math
import pandas as pd
from fastapi import FastAPI, HTTPException, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from pydantic import BaseModel
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from io import BytesIO
from functools import lru_cache
from fastapi.staticfiles import StaticFiles

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:8000",
        "https://efbd-67-173-101-35.ngrok-free.app"
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["*"]
)

def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

def fix_duplicate_observaciones(df):
    cols = df.columns.tolist()
    if "OBSERVACIONES" in cols and "OBSERVACIONES\n" in cols:
        df = df.drop(columns=["OBSERVACIONES"])
    return df

def normalize_dataframe(df, required_cols):
    df = fix_duplicate_observaciones(df)
    for col in required_cols:
        if col not in df.columns:
            df[col] = ""
    df = df[required_cols]
    return df

# ==========================
# Load Maestro Files
# ==========================
try:
    proc_df = pd.read_excel(resource_path("maestro_procedimientos.xlsx"))
    med_df = pd.read_excel(resource_path("maestro_medicamentos.xlsx"))
    diag_df = pd.read_excel(resource_path("maestro_diagnosticos.xlsx"))
except Exception as e:
    raise HTTPException(status_code=500, detail=f"Error loading maestro files: {e}")

# Build medication 'concat' field using real column names:
med_df["concat"] = (
    med_df["CÓDIGO"].astype(str) + " - " +
    med_df["DESCRIPCIÓN"].astype(str) + " " +
    med_df["PRINCIPIO ACTIVO"].astype(str) + " " +
    med_df["FORMA FARMACEUTICA"].astype(str) + " " +
    med_df["CONCENTRACION"].astype(str) + " " +
    med_df["PRESENTACION"].astype(str) + " " +
    med_df["VIA ADMINISTRACION"].astype(str)
)

DATA_FILE = "data.xlsx"

REQUIRED_COLUMNS = [
    'CÓDIGO DEPENDENCIA\n(ESPECIALIDAD)\n',
    'PLANILLA',
    'FECHA ANTENCION',
    'TIPO DE BENEFICIARIO',
    'CEDULA',
    'NOMBRE DE BENEFICIARIO',
    'SEXO-GENERO',
    'FECHA DE NACIMIENTO BENEFICIERO',
    'EDAD BENEFICIERO',
    'TIPO DE SERVICIO/ATENCION',
    'CODIGO',
    'DESCRIPCIÓN',
    'DIAGNOSTICO PRINCIPAL CIE-10',
    'DIAGNSITICO SECUNDARIO 1',
    'DIAGNSITICO SECUNDARIO 2',
    'CANTIDAD',
    'VALOR UNITARIO',
    'DURACION CONSULTA',
    'PARENTESCO',
    'IDENTIFICACION AFILIADO',
    'NOMBRE AFIALIADO',
    'CODIGO DE DERIVACION',
    'NUMERO SECUNCIAL DERIVACION',
    'CONTINGENCIA CUBIERTA',
    'DIAGNOSTICO PRESUNTIVO O DIFINITIVO',
    'TIEMPO ANESTESIA',
    'DIAGNSITICO SECUNDARIO 3',
    'DIAGNSITICO SECUNDARIO 4',
    'DIAGNSITICO SECUNDARIO 5',
    'PORCENTAJE IVA',
    'VALOR IVA',
    'VALOR TOTAL',
    'GASTOS DE GESTIÓN (VALOR\nUNITARIO) / MODIFICADORES NO\nGEOGRÁFICOS (VALOR UNITARIO)',
    'FECHA DE INGRESO',
    'FECHA DE EGRESO',
    'MOTIVO DE EGRESO',
    'COBERTURA COMPARTIDA\n',
    'TIPO DE COBERTURA\n',
    'DISCAPACIDAD CERTIFICADA\n',
    'TIPO DE PRESTACIÓN\n',
    'TIPO DE MÉDICO',
    'FECHA AUTORIZADA PARA INICIO DE ATENCIÓN \n',
    'OBSERVACIONES\n'
]

if os.path.exists(DATA_FILE):
    df = pd.read_excel(DATA_FILE)
    df.columns = df.columns.str.strip()
    df = normalize_dataframe(df, REQUIRED_COLUMNS)
else:
    df = pd.DataFrame(columns=REQUIRED_COLUMNS)

grid_columns = [
    'FECHA DE INGRESO',
    'FECHA DE EGRESO',
    'CÓDIGO DEPENDENCIA (ESPECIALIDAD)',
    'FECHA ANTENCION',
    'CEDULA',
    'NOMBRE DE BENEFICIARIO',
    'CODIGO',
    'DESCRIPCIÓN',
    'DIAGNOSTICO PRINCIPAL CIE-10',
    'DIAGNSITICO SECUNDARIO 1',
    'CANTIDAD',
    'DIAGNOSTICO PRESUNTIVO O DIFINITIVO',
    'OBSERVACIONES'
]

color_fills = [
    PatternFill(start_color="FF92D050", end_color="FF92D050", fill_type="darkGrid"),
    PatternFill(start_color="FF00B0F0", end_color="FF00B0F0", fill_type="darkTrellis"),
    PatternFill(start_color="FFFFC000", end_color="FFFFC000", fill_type="lightGrid"),
    PatternFill(start_color="FF7030A0", end_color="FF7030A0", fill_type="lightTrellis"),
    PatternFill(start_color="FF00B050", end_color="FF00B050", fill_type="darkHorizontal"),
    PatternFill(start_color="FFED7D31", end_color="FFED7D31", fill_type="darkVertical"),
    PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="lightHorizontal"),
    PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="lightVertical"),
    PatternFill(start_color="FFBFBFBF", end_color="FFBFBFBF", fill_type="darkDown"),
    PatternFill(start_color="FFFF00FF", end_color="FFFF00FF", fill_type="darkUp")
]

# -----------------------------
# Pydantic Models
# -----------------------------
class EntryItem(BaseModel):
    name: str = ""
    code: str = ""
    quantity: int = 0

class NewEntry(BaseModel):
    paciente: str = ""
    diagnostico_name: str = ""
    diagnostico_code: str = ""
    diagnostico_secundario_name: str = ""
    diagnostico_secundario_code: str = ""
    fecha_ingreso: str = ""
    fecha_egreso: str = ""
    observaciones: str = ""
    procedimientos: list[EntryItem] = []
    medicamentos: list[EntryItem] = []
    insumos: list[EntryItem] = []

class DeleteRows(BaseModel):
    ids: list[int]

# -----------------------------
# Endpoints
# -----------------------------
@app.post("/upload/")
async def upload_file(file: UploadFile = File(...)):
    global df
    try:
        contents = await file.read()
        file_location = f"./{file.filename}"
        with open(file_location, "wb") as f:
            f.write(contents)
        if file.filename.lower().endswith('.csv'):
            temp_df = pd.read_csv(BytesIO(contents))
        else:
            temp_df = pd.read_excel(BytesIO(contents))
        temp_df.columns = temp_df.columns.str.strip()
        temp_df = normalize_dataframe(temp_df, REQUIRED_COLUMNS)
        df = temp_df
        return {"message": "File uploaded and loaded successfully."}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/data/")
def get_data():
    records = df.to_dict(orient="records")
    for idx, record in enumerate(records):
        record["id"] = idx
        for key, value in record.items():
            if isinstance(value, float) and math.isnan(value):
                record[key] = None
    return records

@app.get("/sync/diagnostic/")
def sync_diagnostic(name: str = None, code: str = None):
    if name:
        row = diag_df[diag_df["NOMBRE"].str.lower() == name.lower()]
    elif code:
        row = diag_df[diag_df["CÓDIGO"].astype(str) == str(code)]
    else:
        raise HTTPException(status_code=400, detail="Provide either name or code")
    if row.empty:
        raise HTTPException(status_code=404, detail="Diagnostic not found")
    return {"name": row.iloc[0]["NOMBRE"], "code": row.iloc[0]["CÓDIGO"]}

@app.get("/search/diagnostics/")
def search_diagnostics(query: str):
    results = diag_df[diag_df["NOMBRE"].str.contains(query, case=False, na=False)].head(50)
    return results[["NOMBRE", "CÓDIGO"]].to_dict(orient="records")

@app.get("/search/diagnostics/code/")
def search_diagnostics_code(query: str):
    mask = diag_df["CÓDIGO"].astype(str).str.contains(query, case=False, na=False)
    results = diag_df[mask].head(50)
    return results[["NOMBRE", "CÓDIGO"]].to_dict(orient="records")

@app.get("/search/procedures/")
def search_procedures(query: str):
    results = proc_df[proc_df["DESCRIPCIÓN"].str.contains(query, case=False, na=False)].head(50)
    return results[["DESCRIPCIÓN", "CÓDIGO"]].to_dict(orient="records")

@app.get("/search/medications/")
def search_medications(query: str):
    results = med_df[med_df["concat"].str.contains(query, case=False, na=False)].head(50)
    # Force CODIGO to be a string
    out = results[["concat", "CÓDIGO"]].to_dict(orient="records")
    for item in out:
        item["CÓDIGO"] = str(item["CÓDIGO"])
    return out

@lru_cache(maxsize=1)
@app.get("/patients/full/")
def get_patients_full():
    return sorted(df["NOMBRE DE BENEFICIARIO"].dropna().unique().tolist())

@lru_cache(maxsize=1)
@app.get("/medications/full/")
def get_medications_full():
    out = []
    for _, row in med_df.iterrows():
        code_val = row["CÓDIGO"] if pd.notnull(row["CÓDIGO"]) else row["concat"].split(" - ")[0]
        out.append({
            "concat": row["concat"],
            "CODIGO": str(code_val)
        })
    return out

@lru_cache(maxsize=1)
@app.get("/procedures/full/")
def get_procedures_full():
    out = []
    for _, row in proc_df.iterrows():
        out.append({
            "DESCRIPCIÓN": row["DESCRIPCIÓN"],
            "CÓDIGO": str(row["CÓDIGO"])
        })
    return out

@lru_cache(maxsize=1)
@app.get("/diagnostics/full/")
def get_diagnostics_full():
    out = []
    for _, row in diag_df.iterrows():
        out.append({
            "NOMBRE": row["NOMBRE"],
            "CÓDIGO": str(row["CÓDIGO"])
        })
    return out

@app.get("/download/")
def download_file():
    if not os.path.exists(DATA_FILE):
        raise HTTPException(status_code=404, detail="Data file not found")
    return FileResponse(
        path=DATA_FILE,
        filename="data.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.post("/add/")
def add_entry(entry: NewEntry):
    global df
    base_row = {
        "NOMBRE DE BENEFICIARIO": entry.paciente,
        "DIAGNOSTICO PRINCIPAL CIE-10": entry.diagnostico_code,
        "DIAGNOSTICO PRESUNTIVO O DIFINITIVO": entry.diagnostico_name,
        "DIAGNSITICO SECUNDARIO 1": entry.diagnostico_secundario_code,
        "FECHA DE INGRESO": entry.fecha_ingreso,
        "FECHA DE EGRESO": entry.fecha_egreso,
        "OBSERVACIONES": entry.observaciones
    }
    new_entries = []
    for item in entry.procedimientos:
        if item.name:
            row = base_row.copy()
            row["DESCRIPCIÓN"] = item.name
            row["CODIGO"] = item.code
            row["CANTIDAD"] = item.quantity
            new_entries.append(row)
    for item in entry.medicamentos:
        if item.name:
            row = base_row.copy()
            row["DESCRIPCIÓN"] = item.name
            row["CODIGO"] = item.code
            row["CANTIDAD"] = item.quantity
            new_entries.append(row)
    for item in entry.insumos:
        if item.name:
            row = base_row.copy()
            row["DESCRIPCIÓN"] = item.name
            row["CODIGO"] = item.code
            row["CANTIDAD"] = item.quantity
            new_entries.append(row)
    current_rows = df.to_dict(orient="records")
    for new_row in new_entries:
        patient = new_row["NOMBRE DE BENEFICIARIO"]
        insertion_index = None
        for i in range(len(current_rows) - 1, -1, -1):
            if current_rows[i]["NOMBRE DE BENEFICIARIO"] == patient:
                new_row["CÓDIGO DEPENDENCIA\n(ESPECIALIDAD)\n"] = current_rows[i].get("CÓDIGO DEPENDENCIA\n(ESPECIALIDAD)\n", "")
                new_row["FECHA ANTENCION"] = current_rows[i].get("FECHA ANTENCION", "")
                new_row["CEDULA"] = current_rows[i].get("CEDULA", "")
                insertion_index = i + 1
                break
        if insertion_index is None:
            current_rows.append(new_row)
        else:
            current_rows.insert(insertion_index, new_row)
    df = pd.DataFrame(current_rows)
    df.to_excel(DATA_FILE, index=False, columns=df.columns.tolist())
    try:
        wb = load_workbook(DATA_FILE)
        ws = wb.active
        header = [cell.value for cell in ws[1]]
        patient_idx = header.index("NOMBRE DE BENEFICIARIO") + 1
        date_idx = header.index("FECHA ANTENCION") + 1
        current_color_index = -1
        prev_key = None
        for r in range(2, ws.max_row + 1):
            patient_val = ws.cell(row=r, column=patient_idx).value
            date_val = ws.cell(row=r, column=date_idx).value
            key = (date_val, patient_val)
            if key != prev_key:
                current_color_index = (current_color_index + 1) % len(color_fills)
                prev_key = key
            fill_to_use = color_fills[current_color_index]
            for c in range(1, ws.max_column + 1):
                ws.cell(row=r, column=c).fill = fill_to_use
        wb.save(DATA_FILE)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error saving colored file: {e}")
    return {"message": "Entry added successfully!"}

@app.post("/delete/")
def delete_rows(delete_request: DeleteRows):
    global df
    records = df.to_dict(orient="records")
    new_records = [record for idx, record in enumerate(records) if idx not in delete_request.ids]
    df = pd.DataFrame(new_records)
    df.to_excel(DATA_FILE, index=False, columns=df.columns.tolist())
    return {"message": "Filas eliminadas exitosamente."}

@app.post("/save/")
def save_file():
    df.to_excel(DATA_FILE, index=False, columns=df.columns.tolist())
    return {"message": "File saved successfully."}

@app.on_event("shutdown")
def save_state():
    pass

build_path = os.path.join(os.path.dirname(__file__), "frontend", "build")
if os.path.exists(build_path):
    app.mount("/", StaticFiles(directory=build_path, html=True), name="static")
else:
    print("Warning: React build directory not found. Frontend will not be served.")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
